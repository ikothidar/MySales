import csv
import sqlite3
from sqlite3 import Error

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


def month_string_to_number(month_string):
    return ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec'].index(
        month_string.lower()[:3]) + 1


def connect_db(db_file):
    con = None

    try:
        con = sqlite3.connect(db_file)
        return con

    except Error as err:
        print(err)

        if con is not None:
            con.close()


def style_header(sheet):
    bd = Side(style='thin', color="000000")
    for cell in sheet:
        cell.font = Font(color="FFFFFF", bold=True, size=12)
        cell.fill = PatternFill("solid", fgColor="7030A0")
        cell.alignment = Alignment(wrap_text=True)
        cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)


def style_data(sheet):
    bd = Side(style='thin', color="000000")
    for cell in sheet:
        cell.alignment = Alignment()
        cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)


def style_total(sheet):
    bd = Side(style='thin', color="000000")
    for cell in sheet:
        cell.alignment = Alignment(wrap_text=True)
        cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        cell.fill = PatternFill("solid", fgColor="B1A0C7")


def primary_sale(sheet, cur, month, year):
    sheet['D1'] = '"Details of Purchase During the Month of {0} {1}"'.format(month, year)
    sheet['D1'].font = Font(color="00B050", bold=True, size=18)
    sheet.merge_cells('D1:J1')

    sheet.append(('',))

    header = ('Date', 'Name of the party', 'GST no of party', 'Invoice No', 'Date', 'Details of Goods', 'HSN code',
              'Total amount of bill', 'Taxable amount before GST', 'IGST @12%', 'IGST @5%', 'CGST @6% and 2.5%',
              'SGST @6% and 2.5%')
    sheet.append(header)

    rows = cur.execute(
        "SELECT * FROM primarysales WHERE strftime('%m', entry_date) = ? and"
        " strftime('%Y', entry_date) = ? ORDER BY entry_date",
        (str(month_string_to_number(month)).zfill(2), str(year)))

    for row in rows:
        sheet.append(row[1:])
        style_data(sheet[sheet.max_row])


def secondary_sale(sheet, cur, month, year):
    sheet['D1'] = '"Details of Sales During the Month of {0} {1}"'.format(month, year)
    sheet['D1'].font = Font(color="00B050", bold=True, size=18)
    sheet.merge_cells('D1:J1')

    sheet.append(('',))

    header = ('Date', 'Name of the party', 'GST no of party', 'Invoice No', 'Date', 'Details of Goods',
              'Total Amount of Bill', 'Taxable Value @12% Ayurvedic Medicine',
              'Taxable Value @12% Compression Germents', 'Taxable Value @5% Abd.Belts', 'CGST', 'SGST',
              'Total GST')

    sheet.append(header)

    rows = cur.execute(
        "SELECT * FROM secondarysales WHERE strftime('%m', entry_date) = ? and "
        "strftime('%Y', entry_date) = ? ORDER BY entry_date",
        (str(month_string_to_number(month)).zfill(2), str(year)))

    for row in rows:
        sheet.append(row[1:])
        style_data(sheet[sheet.max_row])


def sum_amount(sheet, start_col):
    max_row = sheet.max_row + 1
    max_col = sheet.max_column + 1
    sheet['F{}'.format(max_row)] = 'TOTAL'
    for col in range(start_col, max_col):
        total = 0
        for row in range(4, max_row):
            # cell_name = '{0}{1}'.format(col, row)
            total += sheet.cell(row=row, column=col).value

        sheet.cell(column=col, row=max_row).value = total

    style_total(sheet[max_row])


def change_width(sheet):
    for i in range(0, sheet.max_column):
        sheet.column_dimensions[chr(i + 65)].width = 20

    sheet.column_dimensions['B'].width = 30


def create_workbook(month, year, sale_type, db_path):
    mywb = Workbook()

    con = connect_db(db_path)
    cur = con.cursor()

    if sale_type == 'both':
        sheet = mywb.active
        sheet.title = 'Primary'

        primary_sale(sheet, cur, month, year)
        sum_amount(sheet, 8)
        change_width(sheet)
        style_header(sheet[3])

        sheet = mywb.create_sheet('Secondary')

        secondary_sale(sheet, cur, month, year)
        sum_amount(sheet, 7)
        change_width(sheet)
        style_header(sheet[3])

        name = "{0} {1} Purchase & Sale.xlsx".format(month, year)
        mywb.save(r'D:\py\SomeProjects\MySales\Nareshanjali Enterprises\\' + name)

    elif sale_type == 'primary':
        sheet = mywb.active
        sheet.title = 'Primary'

        primary_sale(sheet, cur, month, year)
        sum_amount(sheet, 8)
        change_width(sheet)
        style_header(sheet[3])

        name = '{0} {1} Purchase.xlsx'.format(month, year)
        mywb.save(r'D:\py\SomeProjects\MySales\Nareshanjali Enterprises\\' + name)

    elif sale_type == 'secondary':
        sheet = mywb.active
        sheet.title = 'Secondary'

        secondary_sale(sheet, cur, month, year)
        sum_amount(sheet, 7)
        change_width(sheet)
        style_header(sheet[3])

        name = '{0} {1} Sale.xlsx'.format(month, year)
        mywb.save(r'D:\py\SomeProjects\MySales\Nareshanjali Enterprises\\' + name)

    else:
        print('Please specify the correct sale type of [Primary, Secondary, Both]')

        cur.close()
        con.close()
        mywb.close()


def get_record(db_path, sale_type):
    print('Fetching records from DB...')
    month = input("Enter Full Month Name => ")
    year = input("Enter Year in 'YYYY' format => ")

    create_workbook(month, year, sale_type, db_path)


def insert_record(conn, sql, file_name):
    cur = conn.cursor()
    with open(file_name, 'r') as file:
        records = csv.reader(file)
        cur.executemany(sql, records)
        conn.commit()


def put_record(db_path, sale_type):
    print("Inserting records in DB...")
    conn = connect_db(db_path)
    sql_primary = ''' INSERT INTO primarysales (ENTRY_DATE, NAME_PARTY, GST, INVOICE, DATE, DETAIL_GOODS, HSN_CODE, 
    TOTAL_AMOUNT, TAXABLE_BEFORE_GST, IGST_12per, IGST_5per, CGST, SGST) 
    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?) '''
    sql_secondary = ''' INSERT INTO secondarysales (ENTRY_DATE, NAME_PARTY, GST, INVOICE, DATE, DETAIL_GOODS, 
    TOTAL_AMOUNT, TAXABLE_12PER_AYURVEDIC, TAXABLE_12PER_COMPRESSION, TAXABLE_5PER_ABD_BELTS, CGST, SGST, GST_AMOUNT) 
    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?) '''

    if sale_type == 'both':
        insert_record(conn, sql_primary, 'insert_primary.csv')
        insert_record(conn, sql_secondary, 'insert_secondary.csv')
    elif sale_type == 'primary':
        insert_record(conn, sql_primary, 'insert_primary.csv')
    elif sale_type == 'secondary':
        insert_record(conn, sql_secondary, 'insert_secondary.csv')
    else:
        print('Please specify the correct sale type of [Primary, Secondary, All]')


def main():
    db_path = r'MySales.db'
    sale_type = {1: 'primary', 2: 'secondary', 3: 'both'}

    print("1. Insert new record")
    print("2. Fetch record")
    print("3. Both")
    opt = int(input("Select from above options 1, 2, 3 => "))

    print("1. Primary")
    print("2. Secondary")
    print("3. Both")
    sale_opt = int(input("Select from above options 1, 2, 3 => "))

    if opt == 1:
        put_record(db_path, sale_type[sale_opt])
    elif opt == 2:
        get_record(db_path, sale_type[sale_opt])
    elif opt == 3:
        put_record(db_path, sale_type[sale_opt])
        get_record(db_path, sale_type[sale_opt])
    else:
        print("You pressed wrong options please select from options 1, 2, 3")

    input("Press any key to exit...")


if __name__ == "__main__":
    main()
