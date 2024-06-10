from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# flask module
from flask import send_file

# Local Modules
from mysales.config.const import (
    FetchTypes,
    PRIMARY_SALES_HEADER,
    SECONDARY_SALES_HEADER,
)
from mysales.utils.utils import convert_list_to_strings, format_date


class FetchData:
    """
    Class to Fetch data and create report.
    """

    def __init__(self, start_date, end_date) -> None:
        # self.fetch_type = fetch_type
        self.start_date = start_date
        self.end_date = end_date

    @staticmethod
    def style_header(sheet):
        bd = Side(style='thin', color="000000")
        for cell in sheet:
            cell.font = Font(color="FFFFFF", bold=True, size=12)
            cell.fill = PatternFill("solid", fgColor="7030A0")
            cell.alignment = Alignment(wrap_text=True)
            cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    @staticmethod
    def style_data(sheet):
        bd = Side(style='thin', color="000000")
        for cell in sheet:
            cell.alignment = Alignment()
            cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    @staticmethod
    def style_total(sheet):
        bd = Side(style='thin', color="000000")
        for cell in sheet:
            cell.alignment = Alignment(wrap_text=True)
            cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
            cell.fill = PatternFill("solid", fgColor="B1A0C7")

    @staticmethod
    def sum_amount(sheet, start_col):
        max_row = sheet.max_row + 1
        max_col = sheet.max_column + 1
        sheet['H{}'.format(max_row)] = 'TOTAL'
        for col in range(start_col, max_col):
            total = 0
            for row in range(4, max_row):
                # cell_name = '{0}{1}'.format(col, row)
                total += sheet.cell(row=row, column=col).value

            sheet.cell(column=col, row=max_row).value = total

        FetchData.style_total(sheet[max_row])

    @staticmethod
    def change_width(sheet):
        for i in range(0, sheet.max_column):
            sheet.column_dimensions[chr(i + 65)].width = 20

        sheet.column_dimensions['B'].width = 30

    def primary_sale(self, sheet):
        sheet['D1'] = (
            f'Details of Purchase from "{self.start_date}" to "{self.end_date}"'
        )
        sheet['D1'].font = Font(color="00B050", bold=True, size=18)
        sheet.merge_cells('D1:J1')

        sheet.append(('',))

        sheet.append(PRIMARY_SALES_HEADER)

        from mysales.models.models import PrimarySales

        data = PrimarySales.query.filter(
            PrimarySales.invoice_date >= self.start_date,
            PrimarySales.invoice_date <= self.end_date,
        ).order_by(PrimarySales.invoice_date.asc()).all()

        if data:
            for row in data:
                row_data = ([
                    format_date(row.entry_date),
                    row.party_name,
                    row.gst_number,
                    row.invoice_number,
                    format_date(row.invoice_date),
                    row.goods_details,
                    row.hsn_code,
                    row.gst_percentage,
                    row.total_bill,
                    row.taxable_value,
                    row.igst,
                    row.cgst,
                    row.sgst,
                ])
                sheet.append(row_data)

        FetchData.style_data(sheet[sheet.max_row])

    def secondary_sale(self, sheet):
        sheet['D1'] = (
            f'Details of Sales from "{self.start_date}" to "{self.end_date}"'
        )
        sheet['D1'].font = Font(color="00B050", bold=True, size=18)
        sheet.merge_cells('D1:J1')

        sheet.append(('',))

        sheet.append(SECONDARY_SALES_HEADER)

        from mysales.models.models import SecondarySales

        data = SecondarySales.query.filter(
            SecondarySales.invoice_date >= self.start_date,
            SecondarySales.invoice_date <= self.end_date,
        ).order_by(SecondarySales.invoice_date.asc()).all()

        if data:
            for row in data:
                row_data = ([
                    format_date(row.entry_date),
                    row.party_name,
                    row.gst_number,
                    row.invoice_number,
                    format_date(row.invoice_date),
                    row.goods_details,
                    row.product_type,
                    row.gst_percentage,
                    row.total_bill,
                    row.taxable_value,
                    row.cgst,
                    row.sgst,
                    row.tax_amount,
                ])
                sheet.append(row_data)

        FetchData.style_data(sheet[sheet.max_row])

    def create_workbook(self):
        my_workbook = Workbook()

        # Create Primary Sales sheet
        sheet = my_workbook.active
        sheet.title = FetchTypes.PRIMARY_SALES.value
        self.primary_sale(sheet)
        FetchData.sum_amount(sheet, 9)
        FetchData.change_width(sheet)
        FetchData.style_header(sheet[3])

        # Create Secondary Sales sheet
        sheet = my_workbook.create_sheet(FetchTypes.SECONDARY_SALES.value)
        self.secondary_sale(sheet)
        FetchData.sum_amount(sheet, 9)
        FetchData.change_width(sheet)
        FetchData.style_header(sheet[3])

        # Write workbook to a BytesIO stream
        output = BytesIO()
        my_workbook.save(output)
        output.seek(0)

        # Return the BytesIO stream as a response with appropriate headers
        return send_file(
            output,
            download_name=(
                f'{self.start_date}_{self.end_date}_Purchases_and_Sales.xlsx'
            ),
            as_attachment=True,
            mimetype=(
                'application/vnd.openxmlformats-officedocument.'
                'spreadsheetml.sheet'
            ),
        )
