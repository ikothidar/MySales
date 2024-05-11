from cgi import FieldStorage
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from config.utils import (
    DATABASE_FILENAME,
    DBTables,
    FetchTypes,
    PRIMARY_SALES_HEADER,
    SECONDARY_SALES_HEADER,
    TARGET_LOCATION,
    VALID_MONTHS,
)
from .handler_helper import SQLite


class FetchData:
    """
    Class to Fetch data and create report.
    """
    def __init__(self, form: FieldStorage) -> None:
        self._form = form

    @staticmethod
    def style_header(sheet):
        bd = Side(style='thin', color="000000")
        for cell in sheet:
            cell.font = Font(color="FFFFFF", bold=True, size=12)
            cell.fill = PatternFill("solid", fgColor="7030A0")
            cell.alignment = Alignment(wrap_text=True)
            cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    @staticmethod
    def style_data(sheet):
        bd = Side(style='thin', color="000000")
        for cell in sheet:
            cell.alignment = Alignment()
            cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    @staticmethod
    def style_total(sheet):
        bd = Side(style='thin', color="000000")
        for cell in sheet:
            cell.alignment = Alignment(wrap_text=True)
            cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
            cell.fill = PatternFill("solid", fgColor="B1A0C7")

    @staticmethod
    def sum_amount(sheet, start_col):
        max_row = sheet.max_row + 1
        max_col = sheet.max_column + 1
        sheet['F{}'.format(max_row)] = 'TOTAL'
        for col in range(start_col, max_col):
            total = 0
            for row in range(4, max_row):
                # cell_name = '{0}{1}'.format(col, row)
                total += sheet.cell(row=row, column=col).value

            sheet.cell(column=col, row=max_row).value = total

        FetchData.style_total(sheet[max_row])

    @staticmethod
    def change_width(sheet):
        for i in range(0, sheet.max_column):
            sheet.column_dimensions[chr(i + 65)].width = 20

        sheet.column_dimensions['B'].width = 30



    def primary_sale(self, sheet, month, year):
        sheet['D1'] = '"Details of Purchase During the Month of {0} {1}"'.format(month, year)
        sheet['D1'].font = Font(color="00B050", bold=True, size=18)
        sheet.merge_cells('D1:J1')

        sheet.append(('',))

        sheet.append(PRIMARY_SALES_HEADER)

        db = SQLite(filename=DATABASE_FILENAME, table=DBTables.PRIMARY_SALES.name)
        
        rows = db.get_data(
            condition=(
                f"WHERE strftime('%m', invoice_date) = {str(VALID_MONTHS[month]).zfill(2)}"
                f" and strftime('%Y', invoice_date) = {str(year)}"
            ),
            order_by='invoice_date',
        )

        for row in rows:
            sheet.append(row[1:])
            FetchData.style_data(sheet[sheet.max_row])


    def secondary_sale(self, sheet, month, year):
        sheet['D1'] = '"Details of Sales During the Month of {0} {1}"'.format(month, year)
        sheet['D1'].font = Font(color="00B050", bold=True, size=18)
        sheet.merge_cells('D1:J1')

        sheet.append(('',))

        sheet.append(SECONDARY_SALES_HEADER)

        db = SQLite(filename=DATABASE_FILENAME, table=DBTables.SECONDARY_SALES.name)
        
        rows = db.get_data(
            condition=(
                f"WHERE strftime('%m', invoice_date) = {str(VALID_MONTHS[month]).zfill(2)}"
                f" and strftime('%Y', invoice_date) = {str(year)}"
            ),
            order_by='invoice_date',
        )

        for row in rows:
            sheet.append(row[1:])
            FetchData.style_data(sheet[sheet.max_row])

    def create_workbook(self, month, year, sale_type):
        mywb = Workbook()

        if sale_type == FetchTypes.BOTH_SALES.value:
            sheet = mywb.active
            sheet.title = FetchTypes.PRIMARY_SALES.name

            self.primary_sale(sheet, month, year)
            FetchData.sum_amount(sheet, 8)
            FetchData.change_width(sheet)
            FetchData.style_header(sheet[3])

            sheet = mywb.create_sheet('Secondary')

            self.secondary_sale(sheet, month, year)
            FetchData.sum_amount(sheet, 7)
            FetchData.change_width(sheet)
            FetchData.style_header(sheet[3])

            mywb.save(
                f'{TARGET_LOCATION}{month}_{year}'
                '_Purchase_and_Sales.xlsx'
            )

        elif sale_type == FetchTypes.PRIMARY_SALES.value:
            sheet = mywb.active
            sheet.title = FetchTypes.PRIMARY_SALES.name

            self.primary_sale(sheet, month, year)
            FetchData.sum_amount(sheet, 8)
            FetchData.change_width(sheet)
            FetchData.style_header(sheet[3])

            mywb.save(
                f'{TARGET_LOCATION}{month}_{year}_Purchase.xlsx'
            )

        elif sale_type == FetchTypes.SECONDARY_SALES.value:
            sheet = mywb.active
            sheet.title = FetchTypes.SECONDARY_SALES.name

            self.secondary_sale(sheet, month, year)
            FetchData.sum_amount(sheet, 7)
            FetchData.change_width(sheet)
            FetchData.style_header(sheet[3])

            mywb.save(
                f'{TARGET_LOCATION}{month}_{year}_Sales.xlsx'
            )
        else:
            print('Some error occured please try again...')

            mywb.close()


    def main(self):
        sale_type = self._form.getfirst('sale_type')
        month = self._form.getfirst('month')
        year = self._form.getfirst('year')

        self.create_workbook(month, year, sale_type, DATABASE_FILENAME)

        return {'success': 'Please check the file location.'}
